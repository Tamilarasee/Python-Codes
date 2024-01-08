# Adding a new column D with 90% of C column value with Chart
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path


def processing_workbooks(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1']
    # cell = sheet.cell(1,1)
    # print(cell.value)
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        crct_price = cell.value * 0.9
        crct_price_new_cell = sheet.cell(row, 4)
        crct_price_new_cell.value = crct_price
    # Chart
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)


# get all filenames in current path and pass to function
path = Path()
for file in path.glob('*.xlsx'):
    processing_workbooks(file)
